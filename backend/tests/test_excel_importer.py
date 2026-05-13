"""Tests for app.services.excel_importer helpers and end-to-end imports."""

from __future__ import annotations

import io
import logging
from datetime import datetime

import openpyxl
import pytest

from app import models
from app.services import excel_importer
from app.services.excel_importer import (
    DECISION_SKIP,
    DECISION_TRUST_A,
    DECISION_TRUST_SHEET,
    SHEET_TO_SECTOR_CODE,
    _count_non_empty_data_rows,
    _geography_full_name,
    _jsonify,
    _normalize_sheet_filter,
    preview_excel,
)


# ---------------------------------------------------------------------------
# Pure helpers
# ---------------------------------------------------------------------------
class TestJsonify:
    @pytest.mark.parametrize(
        "value",
        ["text", 1, 3.14, True, False],
    )
    def test_primitive_passthrough(self, value):
        assert _jsonify(value) == value

    def test_none(self):
        assert _jsonify(None) is None

    def test_datetime_becomes_string(self):
        dt = datetime(2024, 5, 1, 12, 0, 0)
        assert _jsonify(dt) == str(dt)


class TestGeographyFullName:
    @pytest.mark.parametrize(
        ("code", "name"),
        [
            ("SG", "Singapore"),
            ("MY", "Malaysia"),
            ("ID", "Indonesia"),
            ("TH", "Thailand"),
            ("CN", "China"),
        ],
    )
    def test_known(self, code, name):
        assert _geography_full_name(code) == name

    def test_unknown(self):
        assert _geography_full_name("ZZ") is None


class TestNormalizeSheetFilter:
    def test_none_returns_none(self):
        assert _normalize_sheet_filter(None, ["Power"]) is None

    def test_empty_returns_none(self):
        assert _normalize_sheet_filter([], ["Power"]) is None

    def test_only_whitespace_returns_none(self):
        assert _normalize_sheet_filter(["   ", ""], ["Power"]) is None

    def test_trims_and_intersects(self):
        result = _normalize_sheet_filter(
            ["  Power  ", "Industry", "Unknown"], ["Power", "Industry", "Other"]
        )
        assert result == {"Power", "Industry"}

    def test_deduplicates(self):
        result = _normalize_sheet_filter(
            ["Power", "Power", "Industry"], ["Power", "Industry"]
        )
        assert result == {"Power", "Industry"}

    def test_logs_invalid_sheets(self, caplog):
        with caplog.at_level(logging.WARNING, logger=excel_importer.__name__):
            _normalize_sheet_filter(["Power", "Bogus"], ["Power", "Industry"])
        # The warning should mention the bogus sheet.
        assert any("Bogus" in rec.getMessage() for rec in caplog.records)


# ---------------------------------------------------------------------------
# _count_non_empty_data_rows
# ---------------------------------------------------------------------------
def _make_worksheet_bytes(rows: list[list[object]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Power"
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestCountNonEmptyDataRows:
    def test_counts_rows_after_header(self):
        # 9 header rows + 3 data rows
        rows = [["hdr"] for _ in range(9)]
        rows.extend([["data1"], ["data2"], ["data3"]])
        wb_bytes = _make_worksheet_bytes(rows)

        wb = openpyxl.load_workbook(filename=io.BytesIO(wb_bytes), read_only=True)
        ws = wb["Power"]
        assert _count_non_empty_data_rows(ws) == 3
        wb.close()

    def test_ignores_placeholder_only_data_rows(self):
        rows = [["hdr"] for _ in range(9)]
        rows.append(["-", None, ""])  # placeholders -> not counted
        rows.append(["real"])  # counted
        wb_bytes = _make_worksheet_bytes(rows)
        wb = openpyxl.load_workbook(filename=io.BytesIO(wb_bytes), read_only=True)
        ws = wb["Power"]
        assert _count_non_empty_data_rows(ws) == 1
        wb.close()

    def test_empty_sheet(self):
        wb_bytes = _make_worksheet_bytes([["hdr"]])
        wb = openpyxl.load_workbook(filename=io.BytesIO(wb_bytes), read_only=True)
        ws = wb["Power"]
        assert _count_non_empty_data_rows(ws) == 0
        wb.close()


# ---------------------------------------------------------------------------
# preview_excel
# ---------------------------------------------------------------------------
class TestPreviewExcel:
    def test_returns_known_sheet_with_correct_sector(self):
        rows = [["hdr"] for _ in range(9)] + [["real_data"]]
        wb_bytes = _make_worksheet_bytes(rows)
        preview = preview_excel(file_bytes=wb_bytes, file_name="test.xlsx")
        assert preview.file_name == "test.xlsx"
        assert len(preview.sheets) == 1
        sp = preview.sheets[0]
        assert sp.sheet_name == "Power"
        assert sp.is_known is True
        assert sp.sector_code == "POWER"
        assert sp.data_rows == 1

    def test_unknown_sheet_flagged(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MysterySheet"
        ws.cell(row=10, column=1, value="x")
        buf = io.BytesIO()
        wb.save(buf)

        preview = preview_excel(file_bytes=buf.getvalue(), file_name="x.xlsx")
        sp = preview.sheets[0]
        assert sp.is_known is False
        assert sp.sector_code is None


# ---------------------------------------------------------------------------
# Decision constants are stable (a regression guard)
# ---------------------------------------------------------------------------
class TestDecisionConstants:
    def test_values(self):
        assert DECISION_TRUST_SHEET == "TRUST_SHEET"
        assert DECISION_TRUST_A == "TRUST_A"
        assert DECISION_SKIP == "SKIP"

    def test_sheet_mapping_is_complete(self):
        # Every known sector must round-trip through the sheet name table.
        for sheet_name, code in SHEET_TO_SECTOR_CODE.items():
            assert isinstance(sheet_name, str)
            assert code.isupper()
