"""Tests for the EcoTEA workbook writer."""

from __future__ import annotations

import math
from pathlib import Path

import openpyxl
import pytest

from app.converters.base_model import MISSING, PowerRecord
from app.converters.ecotea_writer import (
    POWER_COL_ORDER,
    POWER_DATA_START_ROW,
    write_output,
)


def _make_template(path: Path) -> None:
    """Create a minimal EcoTEA-style template at ``path``.

    Rows 1-9 are reserved for headers; the writer should preserve them. Row 9
    is used as the formatting reference (font, etc.).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Power"
    # Row 9 acts as the formatting reference row.
    for col_idx, name in enumerate(POWER_COL_ORDER, start=1):
        ws.cell(row=9, column=col_idx, value=name)
    # Add an existing data row at row 10 to verify clearing.
    ws.cell(row=10, column=1, value="STALE_VALUE")
    wb.save(path)


class TestWriteOutput:
    def test_writes_records_into_template(self, tmp_path):
        template = tmp_path / "tmpl.xlsx"
        output = tmp_path / "out.xlsx"
        _make_template(template)

        records = [
            PowerRecord(
                process_code="PROC1",
                description="First process",
                geography="SG",
                year=2018,
                ef=56.1,
                capex=900,
                commodity="PWRNGA",
            ),
            PowerRecord(
                process_code="PROC2",
                description="Second process",
                geography="SG",
                year=2030,
                ef=94.6,
                capex=1100,
            ),
        ]

        result_path = write_output(records, str(template), str(output), sheet_name="Power")
        assert Path(result_path).exists()

        wb = openpyxl.load_workbook(result_path)
        ws = wb["Power"]

        # The previous data row should have been wiped.
        # POWER_DATA_START_ROW = 10 and we wrote two records, so we expect
        # row 10 to be PROC1 and row 11 to be PROC2.
        process_code_col = POWER_COL_ORDER.index("process_code") + 1
        year_col = POWER_COL_ORDER.index("year") + 1
        assert ws.cell(row=POWER_DATA_START_ROW, column=process_code_col).value == "PROC1"
        assert ws.cell(row=POWER_DATA_START_ROW + 1, column=process_code_col).value == "PROC2"
        assert ws.cell(row=POWER_DATA_START_ROW + 1, column=year_col).value == 2030

    def test_missing_values_written_as_placeholder(self, tmp_path):
        template = tmp_path / "tmpl.xlsx"
        output = tmp_path / "out.xlsx"
        _make_template(template)

        records = [PowerRecord(process_code="P", ef=float("nan"), capex=None)]
        write_output(records, str(template), str(output))

        wb = openpyxl.load_workbook(output)
        ws = wb["Power"]
        ef_col = POWER_COL_ORDER.index("ef") + 1
        capex_col = POWER_COL_ORDER.index("capex") + 1
        # NaN and None should both be normalized to the MISSING placeholder.
        assert ws.cell(row=POWER_DATA_START_ROW, column=ef_col).value == MISSING
        assert ws.cell(row=POWER_DATA_START_ROW, column=capex_col).value == MISSING

    def test_header_rows_preserved(self, tmp_path):
        template = tmp_path / "tmpl.xlsx"
        output = tmp_path / "out.xlsx"
        _make_template(template)

        write_output(
            [PowerRecord(process_code="HELLO")], str(template), str(output)
        )
        wb = openpyxl.load_workbook(output)
        ws = wb["Power"]
        # Row 9 should still contain the column labels from the template.
        assert ws.cell(row=9, column=1).value == POWER_COL_ORDER[0]

    def test_empty_records_clears_data(self, tmp_path):
        template = tmp_path / "tmpl.xlsx"
        output = tmp_path / "out.xlsx"
        _make_template(template)

        write_output([], str(template), str(output))
        wb = openpyxl.load_workbook(output)
        ws = wb["Power"]
        # The stale row 10 cell A must have been cleared.
        assert ws.cell(row=POWER_DATA_START_ROW, column=1).value is None
