"""HTTP-level tests for /api/imports routes."""

from __future__ import annotations

import tempfile
import uuid
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

import openpyxl
from fastapi.testclient import TestClient
from sqlalchemy.exc import SQLAlchemyError


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _build_workbook(sheets: dict[str, list[list[object]]]) -> bytes:
    """Build an EcoTEA-shaped workbook from {sheet_name: rows}.

    Rows are written starting at row 1; the importer treats row 10 and beyond
    as data rows.
    """
    wb = openpyxl.Workbook()
    # Remove the default sheet.
    default = wb.active
    wb.remove(default)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _empty_workbook() -> bytes:
    return _build_workbook({"Power": [["header"]]})


# ---------------------------------------------------------------------------
# /api/imports/preview
# ---------------------------------------------------------------------------
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _inject_conversion_artefact(wb_bytes: bytes) -> str:
    """Store a minimal ConversionArtefact in the convert cache; return its token."""
    from app.routers.convert import ConversionArtefact, _store_artefact

    tmp_dir = Path(tempfile.mkdtemp())
    out_path = tmp_dir / "converted.xlsx"
    out_path.write_bytes(wb_bytes)

    token = "test-" + uuid.uuid4().hex[:12]
    artefact = ConversionArtefact(
        token=token,
        output_path=out_path,
        download_name="converted.xlsx",
        row_count=0,
        sheet_name="Power",
        model_key="VT_SG_PWR",
        source_file_name="src.xlsx",
        template_file_name="template.xlsx",
    )
    _store_artefact(artefact)
    return token


# ---------------------------------------------------------------------------
# /api/imports/preview
# ---------------------------------------------------------------------------
class TestPreviewEndpoint:
    def test_rejects_non_excel(self, client: TestClient):
        response = client.post(
            "/api/imports/preview",
            files={"file": ("foo.txt", b"hello", "text/plain")},
        )
        assert response.status_code == 400

    def test_rejects_empty_file(self, client: TestClient):
        response = client.post(
            "/api/imports/preview",
            files={
                "file": (
                    "empty.xlsx",
                    b"",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 400

    def test_preview_known_sheet(self, client: TestClient):
        rows = [["header"] for _ in range(9)]
        # One actual data row at row 10.
        rows.append(["PROC_CODE"])
        wb_bytes = _build_workbook({"Power": rows, "Unknown Sheet": [["x"]]})

        response = client.post(
            "/api/imports/preview",
            files={
                "file": (
                    "wb.xlsx",
                    wb_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200
        body = response.json()
        assert body["file_name"] == "wb.xlsx"
        sheets = {entry["sheet_name"]: entry for entry in body["sheets"]}
        assert sheets["Power"]["is_known"] is True
        assert sheets["Power"]["sector_code"] == "POWER"
        assert sheets["Power"]["data_rows"] == 1
        assert sheets["Unknown Sheet"]["is_known"] is False

    def test_rejects_oversized_file(self, client: TestClient):
        """File larger than 50 MB must be rejected with 413."""
        big = b"A" * (51 * 1024 * 1024)
        response = client.post(
            "/api/imports/preview",
            files={"file": ("big.xlsx", big, _XLSX_MIME)},
        )
        assert response.status_code == 413

    def test_corrupt_excel_returns_400(self, client: TestClient):
        response = client.post(
            "/api/imports/preview",
            files={
                "file": (
                    "wb.xlsx",
                    b"not an xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 400


# ---------------------------------------------------------------------------
# /api/imports (create)
# ---------------------------------------------------------------------------
class TestCreateImport:
    def test_rejects_non_excel(self, client: TestClient):
        response = client.post(
            "/api/imports",
            files={"file": ("foo.csv", b"a,b,c", "text/csv")},
        )
        assert response.status_code == 400

    def test_rejects_empty(self, client: TestClient):
        response = client.post(
            "/api/imports",
            files={
                "file": (
                    "x.xlsx",
                    b"",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 400

    def test_rejects_oversized_file(self, client: TestClient):
        big = b"B" * (51 * 1024 * 1024)
        response = client.post(
            "/api/imports",
            files={"file": ("big.xlsx", big, _XLSX_MIME)},
        )
        assert response.status_code == 413

    def test_import_success_with_sheets_filter(self, client: TestClient, seeded_sectors):
        """Happy path: valid xlsx + explicit sheet filter → 201."""
        rows = [["header"] for _ in range(9)] + [["PROC_CODE"]]
        wb_bytes = _build_workbook({"Power": rows})
        response = client.post(
            "/api/imports",
            files={"file": ("wb.xlsx", wb_bytes, _XLSX_MIME)},
            data={"sheets": "Power", "imported_by": "pytest", "note": "unit test"},
        )
        assert response.status_code == 201
        body = response.json()
        assert "import_batch_id" in body

    def test_import_success_all_sheets(self, client: TestClient, seeded_sectors):
        """No sheet filter imports all recognized sheets."""
        rows = [["header"] for _ in range(9)] + [["PROC_CODE"]]
        wb_bytes = _build_workbook({"Power": rows})
        response = client.post(
            "/api/imports",
            files={"file": ("wb.xlsx", wb_bytes, _XLSX_MIME)},
        )
        assert response.status_code == 201

    def test_import_sqlalchemy_error_returns_500(self, client: TestClient):
        """SQLAlchemy failure in import_excel → 500."""
        wb_bytes = _build_workbook({"Power": [["header"]]})

        def _raise_sqla(*args, **kwargs):
            raise SQLAlchemyError("simulated db failure")

        with patch("app.routers.imports.import_excel", side_effect=_raise_sqla):
            response = client.post(
                "/api/imports",
                files={"file": ("wb.xlsx", wb_bytes, _XLSX_MIME)},
            )
        assert response.status_code == 500
        assert "Database write failed" in response.json()["detail"]

    def test_import_generic_error_returns_500(self, client: TestClient):
        """Unexpected exception in import_excel → 500."""
        wb_bytes = _build_workbook({"Power": [["header"]]})

        def _raise_generic(*args, **kwargs):
            raise RuntimeError("boom")

        with patch("app.routers.imports.import_excel", side_effect=_raise_generic):
            response = client.post(
                "/api/imports",
                files={"file": ("wb.xlsx", wb_bytes, _XLSX_MIME)},
            )
        assert response.status_code == 500
        assert "Import failed" in response.json()["detail"]


# ---------------------------------------------------------------------------
# /api/imports/conflicts
# ---------------------------------------------------------------------------
class TestConflicts:
    def test_endpoint_returns_shape(self, client: TestClient, seeded_sectors):
        response = client.get("/api/imports/conflicts")
        assert response.status_code == 200
        body = response.json()
        # Response shape is stable regardless of how many other tests have
        # produced conflicts before this one ran.
        assert "total_pending" in body
        assert isinstance(body["groups"], list)
        assert body["total_pending"] == len(
            [row for group in body["groups"] for row in group["rows"]]
        )

    def test_resolve_empty_list_rejected(self, client: TestClient):
        response = client.post("/api/imports/conflicts/resolve", json=[])
        assert response.status_code == 400


# ---------------------------------------------------------------------------
# /api/imports/preview/from-conversion
# ---------------------------------------------------------------------------
class TestPreviewFromConversion:
    def test_unknown_token_404(self, client: TestClient):
        response = client.post(
            "/api/imports/preview/from-conversion?token=does-not-exist",
        )
        assert response.status_code == 404

    def test_import_from_unknown_token_404(self, client: TestClient):
        response = client.post(
            "/api/imports/from-conversion?token=missing",
        )
        assert response.status_code == 404

    def test_preview_from_conversion_success(self, client: TestClient):
        """A valid artefact token returns the sheet preview."""
        rows = [["header"] for _ in range(9)] + [["PROC_CODE"]]
        wb_bytes = _build_workbook({"Power": rows})
        token = _inject_conversion_artefact(wb_bytes)

        response = client.post(f"/api/imports/preview/from-conversion?token={token}")
        assert response.status_code == 200
        body = response.json()
        assert body["file_name"] == "converted.xlsx"
        assert any(s["sheet_name"] == "Power" for s in body["sheets"])

    def test_preview_from_conversion_corrupt_file_returns_400(self, client: TestClient):
        """Artefact contains unreadable bytes → 400."""
        token = _inject_conversion_artefact(b"not-a-workbook")

        response = client.post(f"/api/imports/preview/from-conversion?token={token}")
        assert response.status_code == 400
        assert "Failed to read" in response.json()["detail"]

    def test_import_from_conversion_success(self, client: TestClient, seeded_sectors):
        """Happy-path: valid artefact token + optional metadata → 201."""
        rows = [["header"] for _ in range(9)] + [["PROC_CODE"]]
        wb_bytes = _build_workbook({"Power": rows})
        token = _inject_conversion_artefact(wb_bytes)

        response = client.post(
            f"/api/imports/from-conversion?token={token}",
            data={"sheets": "Power", "imported_by": "pytest", "note": "conversion test"},
        )
        assert response.status_code == 201
        body = response.json()
        assert "import_batch_id" in body

    def test_import_from_conversion_sqlalchemy_error(self, client: TestClient):
        """SQLAlchemy failure during from-conversion import → 500."""
        wb_bytes = _build_workbook({"Power": [["header"]]})
        token = _inject_conversion_artefact(wb_bytes)

        def _raise_sqla(*args, **kwargs):
            raise SQLAlchemyError("db failure")

        with patch("app.routers.imports.import_excel", side_effect=_raise_sqla):
            response = client.post(f"/api/imports/from-conversion?token={token}")
        assert response.status_code == 500
        assert "Database write failed" in response.json()["detail"]

    def test_import_from_conversion_generic_error(self, client: TestClient):
        """Unexpected exception during from-conversion import → 500."""
        wb_bytes = _build_workbook({"Power": [["header"]]})
        token = _inject_conversion_artefact(wb_bytes)

        def _raise_generic(*args, **kwargs):
            raise RuntimeError("unexpected")

        with patch("app.routers.imports.import_excel", side_effect=_raise_generic):
            response = client.post(f"/api/imports/from-conversion?token={token}")
        assert response.status_code == 500
        assert "Import failed" in response.json()["detail"]


# ---------------------------------------------------------------------------
# /api/imports/conflicts/resolve (error branch)
# ---------------------------------------------------------------------------
class TestResolveConflictsErrors:
    def test_resolve_sqlalchemy_error_returns_500(self, client: TestClient):
        """SQLAlchemy failure in resolve_pending_conflicts → 500."""
        resolutions = [{"raw_row_id": 9999, "decision": "SKIP"}]

        def _raise_sqla(*args, **kwargs):
            raise SQLAlchemyError("db failure")

        with patch("app.routers.imports.resolve_pending_conflicts", side_effect=_raise_sqla):
            response = client.post("/api/imports/conflicts/resolve", json=resolutions)
        assert response.status_code == 500
        assert "Database write failed" in response.json()["detail"]
