"""End-to-end import tests for app.services.excel_importer.

These tests build a real EcoTEA-shaped workbook in memory and run it through
the full import pipeline against the SQLite test database.
"""

from __future__ import annotations

import io
from decimal import Decimal
from typing import Any

import openpyxl

from app import models
from app.services.excel_importer import import_excel

# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------
# EcoTEA cell columns A:AL  (rows 1-9 reserved for headers, data starts row 10)
COLUMNS = [
    "A",
    "B",
    "C",
    "D",
    "E",
    "F",
    "G",
    "H",
    "I",
    "J",
    "K",
    "L",
    "M",
    "N",
    "O",
    "P",
    "Q",
    "R",
    "S",
    "T",
    "U",
    "V",
    "W",
    "X",
    "Y",
    "Z",
    "AA",
    "AB",
    "AC",
    "AD",
    "AE",
    "AF",
    "AG",
    "AH",
    "AI",
    "AJ",
    "AK",
    "AL",
]


def _col_to_index(col: str) -> int:
    # Convert spreadsheet letter (A..AL) to 1-based column index.
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _build_workbook(sheets: dict[str, list[dict[str, Any]]]) -> bytes:
    """Build a workbook where each sheet has 9 header rows + rows from row 10."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        # Write 9 dummy header rows.
        for r in range(1, 10):
            ws.cell(row=r, column=1, value=f"hdr{r}")
        # Write data rows starting at row 10.
        for offset, row_dict in enumerate(rows):
            r = 10 + offset
            for col, value in row_dict.items():
                ws.cell(row=r, column=_col_to_index(col), value=value)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------
class TestImportExcelHappyPath:
    def test_single_row_import_writes_all_satellites(self, db_session, seeded_sectors):
        rows = [
            {
                "A": "Power",
                "B": "ESI",
                "C": "WP1",
                "D": "GREF",
                "E": "VT_SG_PWR",
                "F": "WP1",
                "G": "Scenario analysis",
                "H": "PWRNGACCF01",
                "I": "Combined cycle gas turbine",
                "J": "SG",
                "K": 2020,
                "L": 2018,
                "M": 25,
                "N": "N/A",
                "O": 56.1,  # emission_factor
                "P": "PJ",
                "Q": "MSGD2016",
                "R": 900,  # capex
                "S": "GW",
                "T": 30,  # fixed_opex
                "U": "GW*yr",
                "V": 0.5,  # variable_opex
                "W": "PJ",
                "X": "-",
                "Y": "-",
                "Z": 0.5,  # efficiency
                "AA": 0.42,  # tech_efficiency
                "AB": 1,
                "AC": "PWRNGA",
                "AD": "-",
                "AE": "-",
                "AF": 0.85,  # afa
                "AG": 6.0,  # heat_rate
                "AH": 100,  # capacity
                "AI": "UP",  # bound_type
            }
        ]
        wb_bytes = _build_workbook({"Power": rows})

        result = import_excel(
            db_session,
            file_bytes=wb_bytes,
            file_name="test.xlsx",
            imported_by="pytest",
            note="end-to-end smoke test",
        )

        assert result.rows_imported == 1
        assert result.rows_skipped == 0
        assert result.issues == 0
        assert len(result.sheets) == 1

        # Verify satellite rows landed.
        tech = (
            db_session.query(models.TechnologyProcess)
            .filter_by(technology_code="PWRNGACCF01")
            .one()
        )
        assert tech.technology_lifetime_years == 25

        year = (
            db_session.query(models.TechnologyYear)
            .filter_by(technology_id=tech.technology_id, data_year=2020)
            .one()
        )
        assert year is not None

        eco = (
            db_session.query(models.TechnologyYearEcoteaParameter)
            .filter_by(technology_year_id=year.technology_year_id)
            .one()
        )
        # Emission factor and capex are persisted as Decimal in PostgreSQL Numeric,
        # so compare numerically.
        assert Decimal(str(eco.emission_factor)) == Decimal("56.1")
        assert Decimal(str(eco.capex)) == Decimal("900")

    def test_excel_error_recorded_as_issue(self, db_session, seeded_sectors):
        rows = [
            {
                "A": "Power",
                "H": "TECH1",
                "I": "Test",
                "J": "SG",
                "K": 2020,
                "L": 2018,
                "M": 10,
                "O": "#VALUE!",  # emission_factor with formula error
                "AC": "ELC",
            }
        ]
        wb_bytes = _build_workbook({"Power": rows})
        result = import_excel(db_session, file_bytes=wb_bytes, file_name="errors.xlsx")
        assert result.rows_imported == 1
        assert result.issues >= 1

        issues = db_session.query(models.DataQualityIssue).all()
        assert any(i.original_value == "#VALUE!" for i in issues)

    def test_sector_conflict_marks_pending(self, db_session, seeded_sectors):
        # Sheet says Power, column A says Industry — should mark conflict and skip business write.
        rows = [
            {
                "A": "Industry",  # conflicts with sheet name "Power"
                "H": "TECH_CONFLICT",
                "I": "Conflict row",
                "J": "SG",
                "K": 2020,
            }
        ]
        wb_bytes = _build_workbook({"Power": rows})
        result = import_excel(db_session, file_bytes=wb_bytes, file_name="conflict.xlsx")
        assert result.rows_pending == 1
        assert result.rows_imported == 0
        # No technology_process row should have been written.
        assert (
            db_session.query(models.TechnologyProcess)
            .filter_by(technology_code="TECH_CONFLICT")
            .count()
            == 0
        )

    def test_empty_rows_skipped(self, db_session, seeded_sectors):
        rows = [
            {},  # totally empty
            {"H": "TECH2", "I": "real", "J": "SG", "K": 2020, "AC": "ELC"},
        ]
        wb_bytes = _build_workbook({"Power": rows})
        result = import_excel(db_session, file_bytes=wb_bytes, file_name="mixed.xlsx")
        assert result.rows_imported == 1
        assert result.rows_skipped >= 1

    def test_selected_sheets_filter(self, db_session, seeded_sectors):
        rows = [
            {"H": "P_TECH", "I": "Power tech", "J": "SG", "K": 2020, "AC": "ELC"},
        ]
        wb_bytes = _build_workbook(
            {
                "Power": rows,
                "Industry": [{"H": "I_TECH", "I": "Ind tech", "J": "SG", "K": 2020}],
            }
        )

        result = import_excel(
            db_session,
            file_bytes=wb_bytes,
            file_name="x.xlsx",
            selected_sheets=["Power"],
        )
        # Only the Power sheet was imported.
        assert len(result.sheets) == 1
        assert result.sheets[0].sheet_name == "Power"

    def test_unknown_sheet_filter_imports_nothing(self, db_session, seeded_sectors):
        rows = [
            {"H": "X", "I": "x", "J": "SG", "K": 2020, "AC": "ELC"},
        ]
        wb_bytes = _build_workbook({"Power": rows})
        result = import_excel(
            db_session,
            file_bytes=wb_bytes,
            file_name="x.xlsx",
            selected_sheets=["DoesNotExist"],
        )
        # No sheet matched the allowlist intersection with known sectors, so nothing was imported.
        assert result.rows_imported == 0

    def test_reimport_updates_existing_year(self, db_session, seeded_sectors):
        rows = [
            {
                "H": "REUSE_TECH",
                "I": "first",
                "J": "SG",
                "K": 2020,
                "O": 10,
                "AC": "ELC",
            }
        ]
        wb_bytes = _build_workbook({"Power": rows})
        import_excel(db_session, file_bytes=wb_bytes, file_name="r1.xlsx")

        # Now reimport with a different EF value.
        rows[0]["O"] = 99
        wb_bytes2 = _build_workbook({"Power": rows})
        import_excel(db_session, file_bytes=wb_bytes2, file_name="r2.xlsx")

        # Only one technology_year for (REUSE_TECH, 2020) and EF should be the latest.
        tech = (
            db_session.query(models.TechnologyProcess).filter_by(technology_code="REUSE_TECH").one()
        )
        years = (
            db_session.query(models.TechnologyYear)
            .filter_by(technology_id=tech.technology_id)
            .all()
        )
        assert len(years) == 1
        eco = (
            db_session.query(models.TechnologyYearEcoteaParameter)
            .filter_by(technology_year_id=years[0].technology_year_id)
            .one()
        )
        assert Decimal(str(eco.emission_factor)) == Decimal("99")
