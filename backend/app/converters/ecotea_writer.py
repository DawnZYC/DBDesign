"""Writes ``PowerRecord`` lists into an EcoTEA Excel template.

If the template format changes, update :data:`POWER_COL_ORDER` and the header
rows preserved at the top of :func:`write_output`.
"""

from __future__ import annotations

import shutil

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.converters.base_model import MISSING, PowerRecord


POWER_COL_ORDER = [
    "wp6_title",           # A
    "data_owner",          # B
    "data_provider",       # C
    "data_source",         # D
    "data_source_desc",    # E
    "data_user",           # F
    "usage_purpose",       # G
    "process_code",        # H
    "description",         # I
    "geography",           # J
    "year",                # K
    "start_year",          # L
    "lifetime",            # M
    "grade",               # N
    "ef",                  # O
    "ef_unit",             # P
    "currency",            # Q
    "capex",               # R
    "capex_unit",          # S
    "fixed_opex",          # T
    "fixed_opex_unit",     # U
    "variable_opex",       # V
    "variable_opex_unit",  # W
    "tax_cost",            # X
    "sub_cost",            # Y
    "efficiency",          # Z
    "tech_efficiency",     # AA
    "commodity_share",     # AB
    "commodity",           # AC
    "commodity_demand",    # AD
    "interpolation_rule",  # AE
    "afa",                 # AF
    "heat_rate",           # AG
    "capacity",            # AH
    "capacity_type",       # AI
    "constraint",          # AJ
    "uc_rhsrt",            # AK
    "uc_rhsrt_note",       # AL
]

# Rows 1-9 are header/metadata rows preserved from the template.
POWER_DATA_START_ROW = 10


def write_output(
    records: list[PowerRecord],
    template_path: str,
    output_path: str,
    sheet_name: str = "Power",
) -> str:
    """Write records into the EcoTEA template, preserving all header rows."""
    shutil.copy2(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb[sheet_name]

    ref_row = POWER_DATA_START_ROW - 1

    max_row = ws.max_row
    if max_row >= POWER_DATA_START_ROW:
        for row in ws.iter_rows(min_row=POWER_DATA_START_ROW, max_row=max_row):
            for cell in row:
                cell.value = None

    for row_idx, rec in enumerate(records, start=POWER_DATA_START_ROW):
        for col_idx, field_name in enumerate(POWER_COL_ORDER, start=1):
            val = getattr(rec, field_name, MISSING)
            if val is None or (isinstance(val, float) and str(val) == "nan"):
                val = MISSING

            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val

            ref_cell = ws.cell(row=ref_row, column=col_idx)
            if ref_cell.font:
                cell.font = Font(name=ref_cell.font.name, size=ref_cell.font.size)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(output_path)
    return output_path
