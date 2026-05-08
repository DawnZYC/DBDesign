"""Excel to PostgreSQL import service.

Processing flow for each sheet data row (row >= 10):
  1. Write raw_excel_row.
  2. Map sector by sheet name.
  3. Upsert geography / data_source / commodity dictionaries.
  4. Write traceability_record with A:G plus sheet metadata.
  5. Upsert technology_process from H/I/L/M/N, unique by (tech_code, geography).
  6. Write technology_year from K, unique by (technology_id, data_year).
  7. Write satellite tables: ecotea_parameter / wp_descriptor / commodity / constraint / constraint_detail.
  8. Collect formula errors such as #VALUE! into data_quality_issue.
"""
from __future__ import annotations

import hashlib
import logging
import time
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app import models
from app.schemas import (
    ConflictGroup,
    ConflictListResponse,
    ConflictResolution,
    ConflictResolveResponse,
    ConflictRow,
    FilePreview,
    ImportResult,
    ImportSheetSummary,
    SheetPreview,
)
from app.services.value_cleaner import (
    clean_numeric,
    clean_text,
    is_excel_error,
    is_placeholder,
    parse_commodity_combo,
    parse_efficiency,
    resolve_sector_from_text,
)

logger = logging.getLogger(__name__)

# -----------------------------------------------------------------------------
# Constants
# -----------------------------------------------------------------------------
DATA_START_ROW = 10  # Data starts at row 10; rows 1-9 are headers/metadata.

# raw_excel_row.normalized_status values.
STATUS_NORMALIZED = "normalized"          # Successfully written to business tables.
STATUS_PENDING_SECTOR = "pending_sector_review"  # Sector conflict awaiting user review.
STATUS_SKIPPED = "skipped"                # User chose to skip.
STATUS_PENDING = "pending"                # Initial state.

ISSUE_SECTOR_CONFLICT = "sector_conflict"

# Decision values.
DECISION_TRUST_SHEET = "TRUST_SHEET"
DECISION_TRUST_A = "TRUST_A"
DECISION_SKIP = "SKIP"
VALID_DECISIONS = {DECISION_TRUST_SHEET, DECISION_TRUST_A, DECISION_SKIP}

# Sheet name to sector_code mapping. Sheet name is authoritative; column A is kept as raw backup.
SHEET_TO_SECTOR_CODE: dict[str, str] = {
    "Power": "POWER",
    "Industry": "INDUSTRY",
    "Primary": "PRIMARY",
    "Transport": "TRANSPORT",
    "Water": "WATER",
    "Waste": "WASTE",
    "Building": "BUILDING",
    "Household": "HOUSEHOLD",
    "Agri": "AGRI",
    "InfoComm": "INFOCOMM",
}

# AJ:AL detail columns used to build constraint_detail rows.
CONSTRAINT_DETAIL_COLUMNS: tuple[tuple[str, str], ...] = (
    ("AJ", "max_import_possible"),
    ("AK", "max_solar_output_allowed"),
    ("AL", "capacity_special"),
)


# -----------------------------------------------------------------------------
# Entry point
# -----------------------------------------------------------------------------
def import_excel(
    db: Session,
    *,
    file_bytes: bytes,
    file_name: str,
    imported_by: str | None = None,
    note: str | None = None,
    selected_sheets: list[str] | None = None,
) -> ImportResult:
    """Run one Excel import and return a summary.

    Args:
        selected_sheets: Import only these sheets by sheet name. None or empty imports
            all known sheets. Unknown sheet names are ignored and logged.
    """

    started = time.perf_counter()

    # 1) Create import_batch.
    batch = models.ImportBatch(
        file_name=file_name,
        file_hash=hashlib.sha256(file_bytes).hexdigest(),
        imported_by=imported_by,
        note=note,
        imported_at=datetime.now(timezone.utc),
    )
    db.add(batch)
    db.flush()  # Obtain batch.import_batch_id.

    # 2) Load workbook in read-only data-only mode to avoid formula recalculation.
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=False)

    # 3) Decide which sheets to process from the allowlist.
    requested = _normalize_sheet_filter(selected_sheets, workbook.sheetnames)

    sheet_summaries: list[ImportSheetSummary] = []
    total_imported = 0
    total_skipped = 0
    total_pending = 0
    total_issues = 0

    for sheet_name in workbook.sheetnames:
        if sheet_name not in SHEET_TO_SECTOR_CODE:
            logger.warning("Skipping unrecognized sheet: %s", sheet_name)
            continue
        if requested is not None and sheet_name not in requested:
            logger.info("Skipping sheet due to allowlist: %s", sheet_name)
            continue

        worksheet = workbook[sheet_name]
        summary = _import_sheet(db, batch=batch, worksheet=worksheet, sheet_name=sheet_name)
        sheet_summaries.append(summary)
        total_imported += summary.rows_imported
        total_skipped += summary.rows_skipped
        total_pending += summary.rows_pending
        total_issues += summary.issues

    db.commit()

    duration_ms = int((time.perf_counter() - started) * 1000)
    return ImportResult(
        import_batch_id=batch.import_batch_id,
        file_name=batch.file_name,
        imported_at=batch.imported_at,
        rows_imported=total_imported,
        rows_skipped=total_skipped,
        rows_pending=total_pending,
        issues=total_issues,
        sheets=sheet_summaries,
        duration_ms=duration_ms,
    )


# -----------------------------------------------------------------------------
# Single-sheet import
# -----------------------------------------------------------------------------
def _import_sheet(
    db: Session,
    *,
    batch: models.ImportBatch,
    worksheet: Worksheet,
    sheet_name: str,
) -> ImportSheetSummary:
    """Process all data rows in one sheet.

    Supports sparse-row mode: some sheets such as Transport / Water / Waste /
    Building / Household / Agri / InfoComm store full metadata in row 10, while
    later rows only fill changing fields such as K (year) and AD (demand). An
    empty H column means the technology from the previous row should be reused.
    This function forward-fills column H plus traceability/master metadata.
    """

    sector = _get_sector_by_sheet_name(db, sheet_name=sheet_name)

    # These columns inherit from the nearest previous row with non-empty H in sparse-row mode.
    INHERIT_COLUMNS = ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "L", "M", "N")

    rows_imported = 0
    rows_skipped = 0
    rows_pending = 0
    issues = 0
    inherited_metadata: dict[str, Any] = {}

    last_row = worksheet.max_row or 0
    for excel_row_number in range(DATA_START_ROW, last_row + 1):
        cells = _read_row_as_dict(worksheet=worksheet, row_number=excel_row_number)

        # Skip fully empty rows.
        if all(is_placeholder(v) for v in cells.values()):
            rows_skipped += 1
            continue

        # Sparse-row handling: if H is empty but K (year) exists, forward-fill metadata columns.
        if is_placeholder(cells.get("H")):
            if not inherited_metadata:
                # No previous metadata to inherit, such as a sparse first row.
                rows_skipped += 1
                continue
            for column in INHERIT_COLUMNS:
                if is_placeholder(cells.get(column)):
                    cells[column] = inherited_metadata.get(column)
        else:
            # This row has H, so refresh the metadata cache with non-empty values.
            inherited_metadata = {
                column: cells.get(column)
                for column in INHERIT_COLUMNS
                if not is_placeholder(cells.get(column))
            }

        # 1) Write raw_excel_row with only values actually present in the row, not inherited values.
        raw_cells_payload = _read_row_as_dict(
            worksheet=worksheet, row_number=excel_row_number
        )
        raw_row = models.RawExcelRow(
            import_batch_id=batch.import_batch_id,
            source_sheet_name=sheet_name,
            excel_row_number=excel_row_number,
            row_type="data",
            raw_cells={k: _jsonify(v) for k, v in raw_cells_payload.items()},
            normalized_status=STATUS_PENDING,  # Updated later based on the processing result.
        )
        db.add(raw_row)
        db.flush()

        # 2) Check sector conflicts using only explicit column A values in this row,
        #    excluding inherited values so continuation rows are not falsely flagged.
        explicit_a_value = raw_cells_payload.get("A")
        a_sector_code = resolve_sector_from_text(explicit_a_value)
        if (
            a_sector_code is not None
            and a_sector_code != sector.sector_code
        ):
            # Conflict: do not write business tables yet; mark pending and record an issue.
            raw_row.normalized_status = STATUS_PENDING_SECTOR
            db.add(
                models.DataQualityIssue(
                    raw_row_id=raw_row.raw_row_id,
                    source_sheet_name=sheet_name,
                    excel_row_number=excel_row_number,
                    excel_column="A",
                    issue_type=ISSUE_SECTOR_CONFLICT,
                    original_value=str(explicit_a_value) if explicit_a_value is not None else None,
                    issue_message=(
                        f"sheet name is {sheet_name} (->{sector.sector_code}), "
                        f"but column A is '{explicit_a_value}' (->{a_sector_code})"
                    ),
                )
            )
            rows_pending += 1
            issues += 1
            continue

        # 3) Write business records using inherited cells.
        new_issues = _process_data_row(
            db=db,
            batch=batch,
            sector=sector,
            raw_row=raw_row,
            cells=cells,
            sheet_name=sheet_name,
            excel_row_number=excel_row_number,
        )
        raw_row.normalized_status = STATUS_NORMALIZED
        rows_imported += 1
        issues += new_issues

    return ImportSheetSummary(
        sheet_name=sheet_name,
        rows_total=max(last_row - DATA_START_ROW + 1, 0),
        rows_imported=rows_imported,
        rows_skipped=rows_skipped,
        rows_pending=rows_pending,
        issues=issues,
    )


# -----------------------------------------------------------------------------
# Single-row processing
# -----------------------------------------------------------------------------
def _process_data_row(
    db: Session,
    *,
    batch: models.ImportBatch,
    sector: models.Sector,
    raw_row: models.RawExcelRow,
    cells: dict[str, Any],
    sheet_name: str,
    excel_row_number: int,
) -> int:
    """Process one data row and return the number of new data_quality_issue rows."""

    new_issues = 0

    # ---- Dictionary upserts. Geography remains a dictionary; data_source is merged into traceability. ----
    geography = _upsert_geography(db, code=clean_text(cells.get("J")) or "UNKNOWN")

    # ---- traceability_record ----
    trace = models.TraceabilityRecord(
        sector_id=sector.sector_id,
        raw_row_id=raw_row.raw_row_id,
        wp_title_raw=clean_text(cells.get("A")),
        data_owner_raw=clean_text(cells.get("B")),
        data_provider_raw=clean_text(cells.get("C")),
        data_user_raw=clean_text(cells.get("F")),
        usage_purpose=clean_text(cells.get("G")),
        data_source_name=clean_text(cells.get("D")),
        data_source_description=clean_text(cells.get("E")),
        source_sheet_name=sheet_name,
        source_excel_row=excel_row_number,
    )
    db.add(trace)
    db.flush()

    # ---- technology_process, upserted by (code, geography) ----
    tech = _upsert_technology_process(
        db,
        sector_id=sector.sector_id,
        geography_id=geography.geography_id,
        cells=cells,
    )

    # ---- technology_year ----
    data_year_result = clean_numeric(cells.get("K"))
    if data_year_result.value is None:
        # No year means no anchor row; record an issue.
        _record_issue(
            db,
            raw_row=raw_row,
            sheet_name=sheet_name,
            excel_row_number=excel_row_number,
            column="K",
            issue_type="missing_year",
            original_value=cells.get("K"),
            issue_message="data_year is required but missing; cannot create technology_year anchor",
        )
        return new_issues + 1

    tech_year = _upsert_technology_year(
        db,
        technology_id=tech.technology_id,
        data_year=int(data_year_result.value),
        traceability_id=trace.traceability_id,
        raw_row_id=raw_row.raw_row_id,
    )

    # ---- satellite 1: ecotea_parameter ----
    new_issues += _insert_ecotea_parameter(
        db,
        tech_year=tech_year,
        cells=cells,
        raw_row=raw_row,
        sheet_name=sheet_name,
        excel_row_number=excel_row_number,
    )

    # ---- satellite 2: wp_descriptor ----
    new_issues += _insert_wp_descriptor(
        db,
        tech_year=tech_year,
        cells=cells,
        raw_row=raw_row,
        sheet_name=sheet_name,
        excel_row_number=excel_row_number,
    )

    # ---- junction: technology_year_commodity, preserving multi-commodity order ----
    _insert_commodities(db, tech_year=tech_year, cells=cells)

    # ---- satellite 3: constraint (AH/AI) ----
    new_issues += _insert_constraint(
        db,
        tech_year=tech_year,
        cells=cells,
        raw_row=raw_row,
        sheet_name=sheet_name,
        excel_row_number=excel_row_number,
    )

    # ---- satellite 4: constraint_detail (AJ/AK/AL) ----
    new_issues += _insert_constraint_details(
        db,
        tech_year=tech_year,
        cells=cells,
        raw_row=raw_row,
        sheet_name=sheet_name,
        excel_row_number=excel_row_number,
    )

    return new_issues


# -----------------------------------------------------------------------------
# Upsert helpers
# -----------------------------------------------------------------------------
def _get_sector_by_sheet_name(db: Session, *, sheet_name: str) -> models.Sector:
    """Map sheet name to sector. Sectors should already exist from schema initialization."""
    code = SHEET_TO_SECTOR_CODE[sheet_name]
    sector = db.scalar(select(models.Sector).where(models.Sector.sector_code == code))
    if not sector:
        # Defensive fallback if seed data is missing.
        sector = models.Sector(sector_code=code, sector_name=sheet_name)
        db.add(sector)
        db.flush()
    return sector


def _upsert_geography(db: Session, *, code: str) -> models.Geography:
    """Upsert by geography_code."""
    existing = db.scalar(
        select(models.Geography).where(models.Geography.geography_code == code)
    )
    if existing:
        return existing
    geo = models.Geography(geography_code=code, geography_name=_geography_full_name(code))
    db.add(geo)
    db.flush()
    return geo


def _upsert_commodity(db: Session, *, code: str) -> models.Commodity:
    """Upsert by commodity_code."""
    existing = db.scalar(
        select(models.Commodity).where(models.Commodity.commodity_code == code)
    )
    if existing:
        return existing
    commodity = models.Commodity(commodity_code=code)
    db.add(commodity)
    db.flush()
    return commodity


def _upsert_technology_process(
    db: Session,
    *,
    sector_id: int,
    geography_id: int,
    cells: dict[str, Any],
) -> models.TechnologyProcess:
    """Upsert by (technology_code, geography_id)."""
    code = clean_text(cells.get("H"))
    if not code:
        raise ValueError("technology_code is required")

    existing = db.scalar(
        select(models.TechnologyProcess).where(
            models.TechnologyProcess.technology_code == code,
            models.TechnologyProcess.geography_id == geography_id,
        )
    )
    if existing:
        return existing

    start_year = clean_numeric(cells.get("L")).value
    lifetime = clean_numeric(cells.get("M")).value
    tech = models.TechnologyProcess(
        sector_id=sector_id,
        geography_id=geography_id,
        technology_code=code,
        technology_description=clean_text(cells.get("I")),
        technology_start_year=int(start_year) if start_year is not None else None,
        technology_lifetime_years=int(lifetime) if lifetime is not None else None,
        grade=clean_text(cells.get("N")),
    )
    db.add(tech)
    db.flush()
    return tech


def _upsert_technology_year(
    db: Session,
    *,
    technology_id: int,
    data_year: int,
    traceability_id: int,
    raw_row_id: int,
) -> models.TechnologyYear:
    """Upsert by (technology_id, data_year).

    When the same (technology, year) is imported again:
      - Reuse the existing technology_year row and keep its primary key to avoid
        breaking foreign-key relationships.
      - Clear all satellite rows for ecotea_parameter / wp_descriptor /
        commodity / constraint / constraint_detail so this import can rewrite them.
        Otherwise UNIQUE constraints such as (technology_year_id, commodity_order)
        would conflict.
      - Refresh traceability and raw_row references to this import.
    """
    existing = db.scalar(
        select(models.TechnologyYear).where(
            models.TechnologyYear.technology_id == technology_id,
            models.TechnologyYear.data_year == data_year,
        )
    )
    if existing:
        _clear_satellite_rows(
            db, technology_year_id=existing.technology_year_id
        )
        existing.traceability_id = traceability_id
        existing.raw_row_id = raw_row_id
        db.flush()
        return existing
    tech_year = models.TechnologyYear(
        technology_id=technology_id,
        data_year=data_year,
        traceability_id=traceability_id,
        raw_row_id=raw_row_id,
    )
    db.add(tech_year)
    db.flush()
    return tech_year


def _clear_satellite_rows(db: Session, *, technology_year_id: int) -> None:
    """Delete all satellite rows under one technology_year so imports are reentrant."""
    for model_cls in (
        models.TechnologyYearEcoteaParameter,
        models.TechnologyYearWpDescriptor,
        models.TechnologyYearCommodity,
        models.TechnologyYearConstraint,
        models.TechnologyYearConstraintDetail,
    ):
        db.execute(
            delete(model_cls).where(
                model_cls.technology_year_id == technology_year_id
            )
        )
    db.flush()


# -----------------------------------------------------------------------------
# Satellite writes
# -----------------------------------------------------------------------------
def _insert_ecotea_parameter(
    db: Session,
    *,
    tech_year: models.TechnologyYear,
    cells: dict[str, Any],
    raw_row: models.RawExcelRow,
    sheet_name: str,
    excel_row_number: int,
) -> int:
    """Columns O:Y. Return the number of new issues."""
    issues = 0

    field_to_column = {
        "emission_factor": "O",
        "emission_factor_unit": "P",
        "base_currency": "Q",
        "capex": "R",
        "capex_unit": "S",
        "fixed_opex": "T",
        "fixed_opex_unit": "U",
        "variable_opex": "V",
        "variable_opex_unit": "W",
        "tax_cost": "X",
        "subsidy_cost": "Y",
    }

    numeric_fields = {
        "emission_factor",
        "capex",
        "fixed_opex",
        "variable_opex",
        "tax_cost",
        "subsidy_cost",
    }

    payload: dict[str, Any] = {"technology_year_id": tech_year.technology_year_id}
    has_value = False

    for field, column in field_to_column.items():
        cell = cells.get(column)
        if field in numeric_fields:
            result = clean_numeric(cell)
            payload[field] = result.value
            if result.excel_error:
                _record_issue(
                    db,
                    raw_row=raw_row,
                    sheet_name=sheet_name,
                    excel_row_number=excel_row_number,
                    column=column,
                    issue_type="formula_error",
                    original_value=cell,
                    issue_message=f"{field} field has formula error {result.excel_error}",
                )
                issues += 1
            if result.value is not None:
                has_value = True
        else:
            text = clean_text(cell)
            payload[field] = text
            if text is not None:
                has_value = True

    if not has_value:
        return issues

    db.add(models.TechnologyYearEcoteaParameter(**payload))
    return issues


def _insert_wp_descriptor(
    db: Session,
    *,
    tech_year: models.TechnologyYear,
    cells: dict[str, Any],
    raw_row: models.RawExcelRow,
    sheet_name: str,
    excel_row_number: int,
) -> int:
    """Columns Z, AA, AF, AG. Return the number of new issues."""
    issues = 0

    eff = parse_efficiency(cells.get("Z"))
    if eff.excel_error:
        _record_issue(
            db,
            raw_row=raw_row,
            sheet_name=sheet_name,
            excel_row_number=excel_row_number,
            column="Z",
            issue_type="formula_error",
            original_value=cells.get("Z"),
            issue_message=f"efficiency field has formula error {eff.excel_error}",
        )
        issues += 1

    tech_eff = clean_numeric(cells.get("AA"))
    afa = clean_numeric(cells.get("AF"))
    heat = clean_numeric(cells.get("AG"))

    for column, result in (("AA", tech_eff), ("AF", afa), ("AG", heat)):
        if result.excel_error:
            _record_issue(
                db,
                raw_row=raw_row,
                sheet_name=sheet_name,
                excel_row_number=excel_row_number,
                column=column,
                issue_type="formula_error",
                original_value=cells.get(column),
                issue_message=f"column {column} has formula error {result.excel_error}",
            )
            issues += 1

    has_value = any(
        v is not None
        for v in (eff.value, eff.text, eff.unit, tech_eff.value, afa.value, heat.value)
    )
    if not has_value:
        return issues

    db.add(
        models.TechnologyYearWpDescriptor(
            technology_year_id=tech_year.technology_year_id,
            efficiency_value=eff.value,
            efficiency_text=eff.text,
            efficiency_unit=eff.unit,
            technology_efficiency=tech_eff.value,
            capacity_to_activity_factor=afa.value,
            heat_rate=heat.value,
        )
    )
    return issues


def _insert_commodities(
    db: Session,
    *,
    tech_year: models.TechnologyYear,
    cells: dict[str, Any],
) -> None:
    """Columns AB:AE."""
    shares = parse_commodity_combo(cells.get("AC"), cells.get("AB"))
    if not shares:
        return

    demand_result = clean_numeric(cells.get("AD"))
    interp_text_raw = cells.get("AE")
    interp_value: float | None
    interp_text: str | None
    if is_placeholder(interp_text_raw):
        interp_value, interp_text = None, None
    else:
        result = clean_numeric(interp_text_raw)
        interp_value = result.value
        interp_text = clean_text(interp_text_raw)

    for order, share in enumerate(shares, start=1):
        commodity = _upsert_commodity(db, code=share.code)
        db.add(
            models.TechnologyYearCommodity(
                technology_year_id=tech_year.technology_year_id,
                commodity_id=commodity.commodity_id,
                commodity_order=order,
                commodity_share_value=share.share_value,
                commodity_share_text=share.share_text,
                commodity_demand_value=demand_result.value if order == 1 else None,
                commodity_demand_text=clean_text(cells.get("AD")) if order == 1 else None,
                interpolation_rule_value=interp_value,
                interpolation_rule_text=interp_text,
            )
        )


def _insert_constraint(
    db: Session,
    *,
    tech_year: models.TechnologyYear,
    cells: dict[str, Any],
    raw_row: models.RawExcelRow,
    sheet_name: str,
    excel_row_number: int,
) -> int:
    """Columns AH:AI. Return the number of new issues."""
    cap = clean_numeric(cells.get("AH"))
    bound = clean_text(cells.get("AI"))

    if cap.excel_error:
        _record_issue(
            db,
            raw_row=raw_row,
            sheet_name=sheet_name,
            excel_row_number=excel_row_number,
            column="AH",
            issue_type="formula_error",
            original_value=cells.get("AH"),
            issue_message=f"capacity field has formula error {cap.excel_error}",
        )

    if cap.value is None and not bound:
        return 1 if cap.excel_error else 0

    db.add(
        models.TechnologyYearConstraint(
            technology_year_id=tech_year.technology_year_id,
            constraint_type="capacity",
            constraint_value=cap.value,
            bound_type=bound,
        )
    )
    return 1 if cap.excel_error else 0


def _insert_constraint_details(
    db: Session,
    *,
    tech_year: models.TechnologyYear,
    cells: dict[str, Any],
    raw_row: models.RawExcelRow,
    sheet_name: str,
    excel_row_number: int,
) -> int:
    """Columns AJ:AL. Insert one row for each non-empty column."""
    issues = 0
    for column, detail_type in CONSTRAINT_DETAIL_COLUMNS:
        result = clean_numeric(cells.get(column))
        if result.excel_error:
            _record_issue(
                db,
                raw_row=raw_row,
                sheet_name=sheet_name,
                excel_row_number=excel_row_number,
                column=column,
                issue_type="formula_error",
                original_value=cells.get(column),
                issue_message=f"column {column} ({detail_type}) has formula error {result.excel_error}",
            )
            issues += 1
        if result.value is None:
            continue
        db.add(
            models.TechnologyYearConstraintDetail(
                technology_year_id=tech_year.technology_year_id,
                detail_type=detail_type,
                detail_value=result.value,
            )
        )
    return issues


# -----------------------------------------------------------------------------
# Utilities
# -----------------------------------------------------------------------------
def _record_issue(
    db: Session,
    *,
    raw_row: models.RawExcelRow,
    sheet_name: str,
    excel_row_number: int,
    column: str,
    issue_type: str,
    original_value: Any,
    issue_message: str,
) -> None:
    """Write a data_quality_issue row."""
    db.add(
        models.DataQualityIssue(
            raw_row_id=raw_row.raw_row_id,
            source_sheet_name=sheet_name,
            excel_row_number=excel_row_number,
            excel_column=column,
            issue_type=issue_type,
            original_value=str(original_value) if original_value is not None else None,
            issue_message=issue_message,
        )
    )


def _read_row_as_dict(*, worksheet: Worksheet, row_number: int) -> dict[str, Any]:
    """Read one row and return a {column_letter: value} dictionary."""
    cells: dict[str, Any] = {}
    for col_index in range(1, (worksheet.max_column or 0) + 1):
        letter = get_column_letter(col_index)
        cells[letter] = worksheet.cell(row=row_number, column=col_index).value
    return cells


def _jsonify(value: Any) -> Any:
    """Convert a cell value into a JSON-serializable value, including datetime handling."""
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _geography_full_name(code: str) -> str | None:
    """Expand common geography codes. Return None for unknown values."""
    return {
        "SG": "Singapore",
        "MY": "Malaysia",
        "ID": "Indonesia",
        "TH": "Thailand",
        "CN": "China",
    }.get(code)


def list_pending_conflicts(db: Session) -> ConflictListResponse:
    """List all sector conflicts pending review, grouped by (sheet, raw column A value)."""
    rows = list(
        db.scalars(
            select(models.RawExcelRow)
            .where(models.RawExcelRow.normalized_status == STATUS_PENDING_SECTOR)
            .order_by(
                models.RawExcelRow.source_sheet_name,
                models.RawExcelRow.excel_row_number,
            )
        )
    )

    if not rows:
        return ConflictListResponse(total_pending=0, groups=[])

    # Extract each raw_row column A value and group by (sheet, a_value).
    grouped: dict[tuple[str, str | None], list[models.RawExcelRow]] = {}
    for row in rows:
        a_value = row.raw_cells.get("A") if isinstance(row.raw_cells, dict) else None
        key = (row.source_sheet_name, str(a_value) if a_value is not None else None)
        grouped.setdefault(key, []).append(row)

    groups: list[ConflictGroup] = []
    for (sheet_name, a_value), members in grouped.items():
        sheet_sector = SHEET_TO_SECTOR_CODE.get(sheet_name)
        a_sector = resolve_sector_from_text(a_value)
        first_row = members[0]
        groups.append(
            ConflictGroup(
                group_id=f"{sheet_name}::{a_value or ''}::{first_row.raw_row_id}",
                sheet_name=sheet_name,
                sheet_sector_code=sheet_sector,
                a_column_value=a_value,
                a_column_sector_code=a_sector,
                rows=[
                    ConflictRow(
                        raw_row_id=r.raw_row_id, excel_row_number=r.excel_row_number
                    )
                    for r in members
                ],
                message=(
                    f"{sheet_name} sheet has {len(members)} conflicting rows: "
                    f"the sheet name maps to {sheet_sector}, "
                    f"but column A is '{a_value}' -> {a_sector or 'unresolved'}"
                ),
            )
        )

    return ConflictListResponse(total_pending=len(rows), groups=groups)


def resolve_pending_conflicts(
    db: Session, *, resolutions: list[ConflictResolution]
) -> ConflictResolveResponse:
    """Process pending rows according to user decisions: write business tables or skip."""
    resolved = 0
    failed = 0
    failure_reasons: list[str] = []

    for item in resolutions:
        if item.decision not in VALID_DECISIONS:
            failed += 1
            failure_reasons.append(
                f"raw_row_id={item.raw_row_id}: unknown decision '{item.decision}'"
            )
            continue
        try:
            _resolve_single_conflict(
                db, raw_row_id=item.raw_row_id, decision=item.decision
            )
            resolved += 1
        except Exception as exc:  # noqa: BLE001
            failed += 1
            failure_reasons.append(f"raw_row_id={item.raw_row_id}: {exc!s}")
            db.rollback()

    if resolved > 0 and failed == 0:
        db.commit()
    elif resolved > 0 and failed > 0:
        # Commit partial success; rolled-back failures have no side effects.
        db.commit()

    return ConflictResolveResponse(
        resolved=resolved, failed=failed, failure_reasons=failure_reasons
    )


def _resolve_single_conflict(
    db: Session, *, raw_row_id: int, decision: str
) -> None:
    """Process one pending row."""
    raw_row = db.get(models.RawExcelRow, raw_row_id)
    if raw_row is None:
        raise ValueError(f"raw_row {raw_row_id} does not exist")
    if raw_row.normalized_status != STATUS_PENDING_SECTOR:
        raise ValueError(
            f"raw_row {raw_row_id} status is '{raw_row.normalized_status}', "
            f"not pending_sector_review; it has already been processed"
        )

    if decision == DECISION_SKIP:
        raw_row.normalized_status = STATUS_SKIPPED
        return

    sheet_name = raw_row.source_sheet_name
    cells: dict[str, Any] = dict(raw_row.raw_cells) if isinstance(raw_row.raw_cells, dict) else {}

    if decision == DECISION_TRUST_SHEET:
        sector_code = SHEET_TO_SECTOR_CODE.get(sheet_name)
    else:  # DECISION_TRUST_A
        sector_code = resolve_sector_from_text(cells.get("A"))

    if not sector_code:
        raise ValueError(
            f"cannot derive sector_code from decision='{decision}' "
            f"(sheet={sheet_name}, A={cells.get('A')!r})"
        )

    sector = db.scalar(
        select(models.Sector).where(models.Sector.sector_code == sector_code)
    )
    if sector is None:
        raise ValueError(f"sector_code '{sector_code}' does not exist in the sector table")

    batch = db.get(models.ImportBatch, raw_row.import_batch_id)
    if batch is None:
        raise ValueError(f"corresponding import_batch {raw_row.import_batch_id} does not exist")

    _process_data_row(
        db=db,
        batch=batch,
        sector=sector,
        raw_row=raw_row,
        cells=cells,
        sheet_name=sheet_name,
        excel_row_number=raw_row.excel_row_number,
    )
    raw_row.normalized_status = STATUS_NORMALIZED


def preview_excel(*, file_bytes: bytes, file_name: str) -> FilePreview:
    """Parse only sheet names and row counts without writing to the database for sheet selection."""
    workbook = load_workbook(
        filename=BytesIO(file_bytes), data_only=True, read_only=True
    )
    previews: list[SheetPreview] = []
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            data_rows = _count_non_empty_data_rows(worksheet)
            previews.append(
                SheetPreview(
                    sheet_name=sheet_name,
                    is_known=sheet_name in SHEET_TO_SECTOR_CODE,
                    sector_code=SHEET_TO_SECTOR_CODE.get(sheet_name),
                    data_rows=data_rows,
                )
            )
    finally:
        workbook.close()
    return FilePreview(file_name=file_name, sheets=previews)


# -----------------------------------------------------------------------------
# Internal helpers
# -----------------------------------------------------------------------------
def _normalize_sheet_filter(
    selected: list[str] | None, available: list[str]
) -> set[str] | None:
    """Normalize the allowlist: trim whitespace, deduplicate, and keep only sheets present in the file.

    Return None to mean no filtering, importing all known sheets.
    """
    if not selected:
        return None
    cleaned = {s.strip() for s in selected if s and s.strip()}
    if not cleaned:
        return None
    available_set = set(available)
    valid = cleaned & available_set
    invalid = cleaned - available_set
    if invalid:
        logger.warning("These sheets from the allowlist do not exist in the file and were ignored: %s", sorted(invalid))
    return valid


def _count_non_empty_data_rows(worksheet: Worksheet) -> int:
    """Count non-empty data rows from row >= 10 in read_only mode."""
    count = 0
    for row_index, row_values in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if row_index < DATA_START_ROW:
            continue
        if any(not is_placeholder(v) for v in row_values):
            count += 1
    return count


__all__ = [
    "import_excel",
    "preview_excel",
    "list_pending_conflicts",
    "resolve_pending_conflicts",
]
