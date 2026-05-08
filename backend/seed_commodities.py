"""Seed the full commodity dictionary from the Commodities sheet in VT_SG_PWR_GREF.xlsx.

Safe to run repeatedly. Rows are upserted by commodity_code; existing metadata is overwritten.

Usage:
    python seed_commodities.py /path/to/VT_SG_PWR_GREF.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

# Make the app package importable.
sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import load_workbook
from sqlalchemy import select

from app.database import SessionLocal
from app import models

# Commodities sheet layout from column B onward. It has two side-by-side blocks:
# NRG on the left and ENV on the right. Each block has 8 columns:
# Csets / CommName / CommDesc / Unit / LimType / CTSLvl / PeakTS / Ctype.
LEFT_COLS = {
    "set": "B",
    "code": "C",
    "desc": "D",
    "unit": "E",
    "lim": "F",
    "cts_lvl": "G",
    "peak_ts": "H",
    "ctype": "I",
}
RIGHT_COLS = {
    "set": "K",
    "code": "L",
    "desc": "M",
    "unit": "N",
    "lim": "O",
    "cts_lvl": "P",
    "peak_ts": "Q",
    "ctype": "R",
}


def _is_data_value(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, str):
        s = v.strip()
        return bool(s) and not s.startswith("*")
    return True


def _cell(sheet, col_letter: str, row: int):
    col_idx = sum((ord(c) - ord("A") + 1) * (26 ** i) for i, c in enumerate(reversed(col_letter)))
    return sheet.cell(row=row, column=col_idx).value


def _read_block(sheet, cols: dict, max_row: int) -> list[dict]:
    out: list[dict] = []
    for row in range(2, max_row + 1):
        code = _cell(sheet, cols["code"], row)
        if not _is_data_value(code):
            continue
        # Skip header rows such as CommName / "Commodity name".
        code_lower = str(code).strip().lower()
        if code_lower in {"commname", "commodity name"}:
            continue
        cs = _cell(sheet, cols["set"], row)
        # The set must be a data value such as NRG/ENV; otherwise treat it as a group title or header.
        if not _is_data_value(cs) or str(cs).strip().lower() in {"csets", "~fi_comm"}:
            continue
        out.append({
            "commodity_code": str(code).strip(),
            "commodity_set": _str_or_none(cs),
            "commodity_description": _str_or_none(_cell(sheet, cols["desc"], row)),
            "unit": _str_or_none(_cell(sheet, cols["unit"], row)),
            "lim_type": _str_or_none(_cell(sheet, cols["lim"], row)),
            "cts_lvl": _str_or_none(_cell(sheet, cols["cts_lvl"], row)),
            "peak_ts": _str_or_none(_cell(sheet, cols["peak_ts"], row)),
            "ctype": _str_or_none(_cell(sheet, cols["ctype"], row)),
        })
    return out


def _str_or_none(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s != "-" else None


def main(xlsx_path: Path) -> None:
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")

    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    if "Commodities" not in wb.sheetnames:
        sys.exit(f"sheet 'Commodities' does not exist; available sheets: {wb.sheetnames}")

    sheet = wb["Commodities"]
    max_row = sheet.max_row or 0

    left = _read_block(sheet, LEFT_COLS, max_row)
    right = _read_block(sheet, RIGHT_COLS, max_row)
    all_rows = {row["commodity_code"]: row for row in left}
    for row in right:
        all_rows[row["commodity_code"]] = row  # Right block wins; codes should not overlap in practice.
    wb.close()

    print(f"Parsed {len(all_rows)} commodities from the Commodities sheet:")
    for code, fields in all_rows.items():
        print(
            f"  {code:<18}  set={fields['commodity_set']:<6}  "
            f"unit={fields['unit'] or '-':<5}  desc={fields['commodity_description']}"
        )

    # Write to the database with upsert semantics.
    db = SessionLocal()
    try:
        upserts_new = 0
        upserts_update = 0
        for code, fields in all_rows.items():
            existing = db.scalar(
                select(models.Commodity).where(models.Commodity.commodity_code == code)
            )
            if existing is None:
                db.add(models.Commodity(**fields))
                upserts_new += 1
            else:
                # Existing row: overwrite metadata without changing commodity_id.
                for key, val in fields.items():
                    if key != "commodity_code":
                        setattr(existing, key, val)
                upserts_update += 1
        db.commit()
        print(
            f"\nWrite complete ✓  inserted {upserts_new} rows, updated {upserts_update} rows."
        )
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("Usage: python seed_commodities.py /path/to/VT_SG_PWR_GREF.xlsx")
    main(Path(sys.argv[1]))
